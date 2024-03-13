#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import json
import urllib
import requests
import csv
import re
import calcOneDay
import getNameNumbers

# Calculate the time and date for end of day calculations

xy = calcOneDay.calcOneDay()
start, end = (xy[0], xy[1])


# In[ ]:


import datetime
from datetime import datetime
import dataFile
import getNameNumbers

#
# Get data from the Tempest database for the new station
#

token = '877f6425-04a5-4f33-86e7-7123b7ef53d9'
protocol = 'https://'
urlSiteDevice = 'swd.weatherflow.com/swd/rest/observations/device/'
urlSiteStation = 'swd.weatherflow.com/swd/rest/observations/station/'
deviceID = '246921'
stationID = '95775'
preToken = '&token='
preStart = '?time_start='
preEnd = '&time_end='
start_time = start
end_time = end
dayOffset = '&day_offset=1'
format1 = '&format=csv'

#
# Put it together
# 

goGetDeviceSummary = (f'{protocol}{urlSiteDevice}{deviceID}{preStart}{start_time}{preEnd}{end_time}{format1}{preToken}{token}')
print(goGetDeviceSummary)
r =  requests.get(goGetDeviceSummary)
path = '/home/ec2-user/'
file_name = 'tempest_temp.csv'
full_file = f'{path}{file_name}'

with open(full_file,'w') as fd:
     fd.write(r.text)


# In[ ]:


import pandas as pd
import dataFile
import getNameNumbers
import sqlalchemy
import mysql.connector
import sqlite3

#
# Read in the CSV file for processing in pandas
#

path = '/home/ec2-user/'
file_name = 'tempest_temp.csv'
full_file = f'{path}{file_name}'
df = pd.read_csv(full_file, index_col=False)

pd.set_option('display.max_rows', 1440)
pd.set_option('display.max_columns', 35)
pd.set_option('display.width', 1500)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 2)

max_temp  = (df.sort_values(by='temperature', ascending=False))
max_T = max_temp.iloc[:1]
maxT = max_T['temperature'].values[0]
maxT = round((maxT*1.8) + 32)

min_temp  = (df.sort_values(by='temperature', ascending=True))
min_T = min_temp.iloc[:1]
minT = min_T['temperature'].values[0]
minT = round((minT*1.8) + 32)

max_pres = (df.sort_values(by='pressure', ascending=False))
max_P = max_pres.iloc[:1]
maxP = max_P['pressure'].values[0]

min_pres = (df.sort_values(by='pressure', ascending=True))
min_P = min_pres.iloc[:1]
minP = min_P['pressure'].values[0]

cor_rain = (df.sort_values(by='local_daily_precip_final', ascending=False))
corR_rain = cor_rain.iloc[:1]
corR = corR_rain['local_daily_precip_final'].values[0]
corR = round((corR*0.03937), 2)

tot_rain = df['precip'].sum()
totR = round((tot_rain*0.03937), 2)
print(totR)

strike_distance = (df.sort_values(by='strike_distance', ascending=False))
lightning1 = (df['strike_distance'].between(1,8))
lightning2 = (df['strike_distance'].between(9,16))

df.insert(10,'lightning1',lightning1)
df.insert(11,'lightning2',lightning2)

#Determine if the strike distance is close enough to count as a thunderstorm
x = len(df)
a = 0

while a < x:
        if (df['lightning1'] == True).any():
            q = "Yes"
        else:
            q = "No"
        if (df['lightning2'] == True).any():   
            r1 = "Yes"
        else:
            r1 = "No"
        a += 1    
  
 
#strike_count = df['strike_count'].sum()

#database_username = 'chuckwx'
#database_password = 'jfr716!!00'
#database_ip       = '3.135.162.69'
#database_name     = 'tempestf6'
#database_connection = sqlalchemy.create_engine('mysql+mysqlconnector://{0}:{1}@{2}/{3}'.
#                                               format(database_username, database_password, 
#                                                      database_ip, database_name), connect_args={'connect_timeout': 30})
#df.to_sql(con=database_connection, name='tempestF6', if_exists='replace')


# In[ ]:


#
# Write to the appropriate Excel file
#

import openpyxl
from openpyxl import load_workbook
import datetime
from datetime import datetime
import excelFilename
import calcTimeNow
from calcTimeNow import calcTimeNow
import getNameNumbers

#
# Create the month name for the xlsx filename
#

gg = getNameNumbers.tempest_ec2()

xls_filename, xls_fullname, path_name, date, thisYear, this_month = gg[0], gg[1], gg[2], gg[3], gg[4], gg[5]


#sa1 = excelFilename.tempest_ec2()
#xls_filename, path_name = (sa1[0], sa1[1])
#xls_fullfile = (f'{path_name}{xls_filename}')
#t1 = calcTimeNow()
#now, myMonth, myYear, date = (t1[0], t1[1], t1[2], t1[3])

wb = openpyxl.load_workbook(xls_fullname)
sheet = wb.active

# Write headers first...
a1 = sheet['A1']
a1.value = "Year"
b1 = sheet['B1']
b1.value = thisYear
c1 = sheet['C1']
c1.value = 'Month'
d1 = sheet['D1']
d1.value = this_month

a3 = sheet['A3']
a3.value = "Date"
b3 = sheet['B3']
b3.value = 'High'
c3 = sheet['C3']
c3.value = 'Low'
d3 = sheet['D3']
d3.value = 'Average'

e3 = sheet['E3']
e3.value = "HDD"
f3 = sheet['F3']
f3.value = 'CDD'
g3 = sheet['G3']
g3.value = 'totR'
h3 = sheet['H3']
h3.value = 'corR'

i3 = sheet['i3']
i3.value = 'Lightning1_5'
j3 = sheet['j3']
j3.value = 'Lightning6_10'

# Write the data..

print(date)
offset_day = (int(date) + 2)
maxTT = sheet.cell(row = offset_day, column = 2)
maxTT.value = maxT
minTT = sheet.cell(row = offset_day, column = 3)
minTT.value = minT

totRR = sheet.cell(row = offset_day, column = 7)
totRR.value = totR

corRR = sheet.cell(row = offset_day, column = 8)
corRR.value = corR

lightning1 = sheet.cell(row = offset_day, column = 9)
lightning1.value = q

lightning2 = sheet.cell(row = offset_day, column = 10)
lightning2.value = r1

wb.save(xls_fullname)


# In[ ]:


# Read the Excel file as a possible pandas dataframe and html file

html_path = '/var/www/html/000/'

df1 = pd.read_excel(xls_fullname, skiprows=2)
print(df1)
df2 = df1.drop(df1.columns[[10,11,12,13,14]], axis = 1)
df2.to_html(f'{html_path}testTempest.html', index = False) 

