#!/usr/bin/env python
# coding: utf-8

# In[1]:


import datetime
import time
from datetime import timezone
from datetime import datetime, timedelta

now = datetime.now()
year = now.strftime("%Y") 
month = now.strftime("%m") 
myMonth = now.strftime("%B")
myYear = now.strftime("%Y")
date = now.strftime("%d") 
hourN = now.strftime("%H")
minN = now.strftime("%M")
its_now = (f'{year},{month},{date},{hourN},{minN}')
its_nower = datetime.strptime(its_now, '%Y,%m,%d,%H,%M')
tNow = int(its_nower.timestamp())
tNow = str(tNow)
add_tNow = ("&t=" + tNow)

tt =  int(time.time())
tt = str(tt)

#end_it = (f'{year},{month},{date},0,0')
#end = datetime.strptime(end_it, '%Y,%m,%d,%H,%M')
#end = int(round(end.timestamp()))
#end = str(end)
end = tt

current_day = datetime.now()
previous_day = datetime.now() - (timedelta(days = 1))
print(previous_day)
prev_year = previous_day.strftime("%Y")
prev_month = previous_day.strftime("%m")
prev_date = previous_day.strftime("%d")
start_it = (f'{prev_year},{prev_month},{prev_date},0,2')
start = datetime.strptime(start_it, '%Y,%m,%d,%H,%M')
start = int(round(start.timestamp()))
start = str(start)
print(start,end)


# In[2]:


import collections
import hashlib
import hmac
import time
from datetime import datetime
import requests
import json
import pprint

parameters = {
  "api-key": "vy8jbrjsxlbwgojepq3vfyfqfywyhvbd", 
  "api-secret": "sdqfm6wdfy9w0pqp2vdka38o6b4vcsvc",
  "station-id": 81211, 
  "end-timestamp": end,
  "start-timestamp": start,
  "t": int(time.time())
}

parameters = collections.OrderedDict(sorted(parameters.items()))

for key in parameters:
    print("Parameter name: \"{}\" has value \"{}\"".format(key, parameters[key]))

apiSecret = parameters["api-secret"];
parameters.pop("api-secret", None);

data = ""
for key in parameters:
    data = data + key + str(parameters[key])

print("Data string to hash is: \"{}\"".format(data))
print('\n')

"""
Calculate the HMAC SHA-256 hash that will be used as the API Signature.
"""
apiSignature = hmac.new(
  apiSecret.encode('utf-8'),
  data.encode('utf-8'),
  hashlib.sha256
).hexdigest()

"""
Let's see what the final API Signature looks like.
"""
print("API Signature is: \"{}\"".format(apiSignature))
print('\n')

# Building the URL to get the station

first_part = ('https://api.weatherlink.com/v2/historic/81211?')
api_key = ('api-key=vy8jbrjsxlbwgojepq3vfyfqfywyhvbd')
add_apisig = ('&api-signature=')
add_t = ('&t='+ str(int(time.time())))

start1 = "&start-timestamp=" + start
end1 = "&end-timestamp=" + end

#
URLfinal = (first_part + api_key + add_t + start1 + end1 + add_apisig + apiSignature)
print(URLfinal)

r =  requests.get(URLfinal)
r.encoding = 'utf-8'
#print (r.json())
#print (r.encoding)
#print (r.headers)
print(r)


r =  requests.get(URLfinal)
data_dir_file = '/Users/jameshayes/davis.json' 
with open(data_dir_file, "w") as fd:   
     json.dump(r.json(), fd)


# In[3]:


import collections
import hashlib
import hmac
import time
from datetime import datetime
import requests
import json
import pprint

with open('/Users/jameshayes/davis.json') as fr:
    davisAPI = json.load(fr)
    
    
a = davisAPI['sensors']    
b = a[1]
c = (b['data'])
cLen = len(c)
print(cLen)

with open('/Users/jameshayes/testout.csv', 'w') as outfile: 
    i = 0
    while i < cLen:
        d = c[i]
        hi_temp = (d['temp_hi'])
        lo_temp = (d['temp_lo'])
        rainfall = (d['rainfall_in'])
        print(f'{hi_temp},{lo_temp},{rainfall}', file = outfile)
        i += 1


# In[4]:


import pandas as pd

#
# Read in the CSV file for processing in pandas
#

path_name = '/Users/jameshayes/'
file_name = 'testout.csv'
full_file = (f'{path_name}{file_name}')

df = pd.read_csv(full_file, index_col=False,names=['temp_hi', 'temp_lo', 'rainfall'])


# In[5]:


pd.set_option('display.max_rows', 1440)
pd.set_option('display.max_columns', 35)
pd.set_option('display.width', 1500)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 2)

max_temp  = (df.sort_values(by='temp_hi', ascending=False))
max_T = max_temp.iloc[:1]
maxT = max_T['temp_hi'].values[0]
maxT = round(maxT)

min_temp  = (df.sort_values(by='temp_lo', ascending=True))
min_T = min_temp.iloc[:1]
minT = min_T['temp_lo'].values[0]
minT = round(minT)

tot_rain = df['rainfall'].sum()
totR = round(tot_rain)
print(maxT,minT,rainfall)


# In[6]:


#
# Write to the appropriate Excel file
#

import openpyxl
from openpyxl import load_workbook

#
# Create the month name for the xlsx filename
#

xls_filename = f'{myMonth}_{myYear}_Davis'
xls_suffix = '.xlsx'
xls_fullfile = (f'{path_name}{xls_filename}{xls_suffix}')

wb = openpyxl.load_workbook(xls_fullfile)
sheet = wb.active

# Write headers first...
a1 = sheet['A1']
a1.value = "Year"
b1 = sheet['B1']
b1.value = myYear
c1 = sheet['C1']
c1.value = 'Month'
d1 = sheet['D1']
d1.value = myMonth

a3 = sheet['A3']
a3.value = "Date"
b3 = sheet['B3']
b3.value = 'High'
c3 = sheet['C3']
c3.value = 'Low'
d3 = sheet['D3']
d3.value = 'Average'

e3 = sheet['E3']
e3.value = "HDD"
f3 = sheet['F3']
f3.value = 'CDD'
g3 = sheet['H3']
g3.value = 'totR'

k3 = sheet['K3']
k3.value = "Highs >=90"
k4 = sheet['K4']
k4.value = "Highs <= 32"
k5 = sheet['K5']
k5.value = 'Lows <= 32'
k6 = sheet['K6']
k6.value = 'Lows <= 0'

k13 = sheet['K13']
k13.value = "Total Rainfall"
k14 = sheet['K14']
k14.value = "rain>=0.01"
k15 = sheet['K15']
k15.value = 'rain>=0.01'
k16 = sheet['K16']
k16.value = 'rain>=0.50'
k17= sheet['K17']
k17.value = 'rain>=1.00'
k23 = sheet['K23']
k23.value = 'Monthly Average'
k24 = sheet['K24']
k24.value = 'Departure'

m3 = sheet['M3']
m3.value = "High"
m4 = sheet['M4']
m4.value = "Low"
m13 = sheet['M13']
m13.value = "Max Rain"
m23 = sheet['M23']
m23.value = "Monthy Rainfall"
m24 = sheet['M24']
m24.value = "Departure"

o3 = sheet['O3']
o3.value = "Date"
o4 = sheet['O4']
o4.value = "Date"


# Write the data...
offset_day = (int(date) + 2)
maxTT = sheet.cell(row = offset_day, column = 2)
maxTT.value = maxT
minTT = sheet.cell(row = offset_day, column = 3)
minTT.value = minT
totRR = sheet.cell(row = offset_day, column = 7)
totRR.value = totR

wb.save(xls_fullfile)


# In[ ]:




