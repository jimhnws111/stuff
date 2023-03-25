#!/usr/bin/env python
# coding: utf-8

# In[146]:


import json
import urllib
import requests
import csv
import re
import calcOneDay

# Calculate the time and date for end of day calculations

xy = calcOneDay.calcOneDay()
start = xy[0]
end = xy[1]


# In[154]:


import collections
import hashlib
import hmac
import time
from datetime import datetime
import requests
import json
import pprint

parameters = {
  "api-key": "vy8jbrjsxlbwgojepq3vfyfqfywyhvbd", 
  "api-secret": "sdqfm6wdfy9w0pqp2vdka38o6b4vcsvc",
  "station-id": 81211, 
  "end-timestamp": end,
  "start-timestamp": start,
  "t": int(time.time())
}

parameters = collections.OrderedDict(sorted(parameters.items()))

for key in parameters:
    print("Parameter name: \"{}\" has value \"{}\"".format(key, parameters[key]))

apiSecret = parameters["api-secret"];
parameters.pop("api-secret", None);

data = ""
for key in parameters:
    data = data + key + str(parameters[key])

print("Data string to hash is: \"{}\"".format(data))
print('\n')

"""
Calculate the HMAC SHA-256 hash that will be used as the API Signature.
"""
apiSignature = hmac.new(
  apiSecret.encode('utf-8'),
  data.encode('utf-8'),
  hashlib.sha256
).hexdigest()

"""
Let's see what the final API Signature looks like.
"""
print("API Signature is: \"{}\"".format(apiSignature))
print('\n')

# Building the URL to get the station

first_part = ('https://api.weatherlink.com/v2/historic/81211?')
api_key = ('api-key=vy8jbrjsxlbwgojepq3vfyfqfywyhvbd')
add_apisig = ('&api-signature=')
add_t = ('&t='+ str(int(time.time())))

start1 = "&start-timestamp=" + start
end1 = "&end-timestamp=" + end

#
URLfinal = (first_part + api_key + add_t + start1 + end1 + add_apisig + apiSignature)
print(URLfinal)

r =  requests.get(URLfinal)
print(r)
data_file = '/home/ec2-user/davis.json' 
with open(data_file, "w") as fd:   
     json.dump(r.json(), fd)


# In[155]:


import time
from datetime import datetime
import requests
import json

with open(data_file) as fr:
    davisAPI = json.load(fr) 
    
a = davisAPI['sensors']    
b = a[1]
c = (b['data'])
cLen = len(c)
print(cLen)

with open('/home/ec2-user/davisOut.csv', 'w') as outfile: 
    i = 0
    while i < cLen:
        d = c[i]
        hi_temp = (d['temp_hi'])
        lo_temp = (d['temp_lo'])
        rainfall = (d['rainfall_in'])
        print(f'{hi_temp},{lo_temp},{rainfall}', file = outfile)
        i += 1


# In[149]:


#
# Read in the CSV file for processing in pandas
#

path_name = '/home/ec2-user/'
file_name = 'davisOut.csv'
full_file = (f'{path_name}{file_name}')


# In[150]:


import pandas as pd

df = pd.read_csv(full_file, index_col=False,names=['temp_hi', 'temp_lo', 'rainfall'])

pd.set_option('display.max_rows', 1440)
pd.set_option('display.max_columns', 35)
pd.set_option('display.width', 1500)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 2)

max_temp  = (df.sort_values(by='temp_hi', ascending=False))
max_T = max_temp.iloc[:1]
maxT = max_T['temp_hi'].values[0]
maxT = round(maxT)

min_temp  = (df.sort_values(by='temp_lo', ascending=True))
min_T = min_temp.iloc[:1]
minT = min_T['temp_lo'].values[0]
minT = round(minT)

totR = df['rainfall'].sum()


# In[151]:


#
# Write to the appropriate Excel file
#

import openpyxl
from openpyxl import load_workbook
import excelFilename

#
# Create the month name for the xlsx filename
#

now = datetime.now()
myMonth = now.strftime("%B")
myYear = now.strftime("%Y")

xls_filename = excelFilename.davis()
xls_fullfile = (f'{path_name}{xls_filename}')

wb = openpyxl.load_workbook(xls_fullfile)
sheet = wb.active

# Write headers first...
a1 = sheet['A1']
a1.value = "Year"
b1 = sheet['B1']
b1.value = myYear
c1 = sheet['C1']
c1.value = 'Month'
d1 = sheet['D1']
d1.value = myMonth

a3 = sheet['A3']
a3.value = "Date"
b3 = sheet['B3']
b3.value = 'High'
c3 = sheet['C3']
c3.value = 'Low'
d3 = sheet['D3']
d3.value = 'Average'

e3 = sheet['E3']
e3.value = "HDD"
f3 = sheet['F3']
f3.value = 'CDD'
g3 = sheet['G3']
g3.value = 'Rainfall'

k4 = sheet['K4']
k4.value = "Highs >=90"
k5 = sheet['K5']
k5.value = "Highs <= 32"
k6 = sheet['K6']
k6.value = 'Lows <= 32'
k7 = sheet['K7']
k7.value = 'Lows <= 0'

k13 = sheet['K14']
k13.value = "Total Rainfall"
k14 = sheet['K15']
k14.value = "rain>=0.01"
k15 = sheet['K16']
k15.value = 'rain>=0.10'
k16 = sheet['K17']
k16.value = 'rain>=0.50'
k17= sheet['K18']
k17.value = 'rain>=1.00'
k23 = sheet['K24']
k23.value = 'Monthly Average'
k24 = sheet['K25']
k24.value = 'Departure'

m3 = sheet['M4']
m3.value = "High"
m4 = sheet['M5']
m4.value = "Low"
m13 = sheet['M14']
m13.value = "Max Rain"
m23 = sheet['M24']
m23.value = "Monthy Rainfall"
m24 = sheet['M25']
m24.value = "Departure"

o3 = sheet['O4']
o3.value = "Date"
o4 = sheet['O5']
o4.value = "Date"


# Calculate the date and write the data...
now = datetime.now()
date = now.strftime("%d") 
offset_day = (int(date) + 2)

maxTT = sheet.cell(row = offset_day, column = 2)
maxTT.value = maxT
minTT = sheet.cell(row = offset_day, column = 3)
minTT.value = minT
totRR = sheet.cell(row = offset_day, column = 7)
totRR.value = totR

wb.save(xls_fullfile)


# In[152]:


# Read the Excel file as a possible pandas dataframe and html file

path_to_file = '/home/ec2-user/'
html_path = '/var/www/html/000/'
new_file = f'{path_to_file}{xls_filename}'

df1 = pd.read_excel(new_file, skiprows = 2, names = ['Date','High','Low','Average','HDD','CDD','Rainfall','dead0','dead1','dead2','dead3','dead4','dead5','dead6','dead7'])
df1 = df1.drop(df1.columns[[7,8,9,10,11,12,13,14]], axis = 1)
df1
df1.to_html(f'{html_path}testDavis.html', index = False) 


# In[ ]:




