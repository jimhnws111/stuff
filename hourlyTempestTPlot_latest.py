#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import pandas as pd
import datetime
from datetime import datetime
from dateutil.tz import tzutc, tzlocal
import pytz

#
# Read in the CSV file for processing in pandas
#

path_to_file = '/home/ec2-user/'
full_file =(f'{path_to_file}latestTempest1m.csv')
df = pd.read_csv(full_file, index_col=False)

pd.set_option('display.max_rows', 1440)
pd.set_option('display.max_columns', 35)
pd.set_option('display.width', 1500)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 2)

latestT = (df['temperature'].iloc[-1]) 
print(latestT)
lastT = round((latestT*1.8) + 32)

timezone = pytz.timezone("America/New_York")
recentT = (df['timestamp'].iloc[-1])
dt_object = datetime.fromtimestamp(recentT)
localT = dt_object.astimezone(timezone)
lastTime = localT.strftime('%I:%M %p')  
time24 = localT.strftime('%-H')  
time24 = int(time24)
print(time24)

print(lastT, lastTime)

# write the data to a csv file with an html suffix

#new_path = '/Users/jameshayes/'
new_path = '/home/ec2-user/'

with open(f'{new_path}LatestT_tempest.csv', 'w') as outfile:
    print(f'{lastT},{lastTime}',file = outfile)


# In[ ]:


import openpyxl
from openpyxl import load_workbook
import datetime
from datetime import datetime
import excelFilename
import calcTimeNow
from calcTimeNow import calcTimeNow
import getNameNumbers_ec1


# In[ ]:


#
# Create the month name for the xlsx filename
#

gg = getNameNumbers_ec1.tempest_ec2()

xls_filename, xls_fullfile, path_name, date, this_month, thisYear = gg[0], gg[1], gg[2], gg[3], gg[4], gg[5]
new_path = '/home/ec2-user/'

#
# Create a new filename based on date info
#

finalFile = f'{new_path}{this_month}{date}{thisYear}_TempestT.xlsx'
print(finalFile)

wb = openpyxl.load_workbook(finalFile)
sheet = wb.active

a3 = sheet['A1']
a3.value = 'Time'
b3 = sheet['B1']
b3.value = 'Temperature'

# Write the data..

nowTime = sheet.cell(row = time24 + 2, column = 1)
nowTime.value = lastTime

newest_T = sheet.cell(row = time24 + 2, column = 2)
newest_T.value = lastT

wb.save(finalFile)


# In[ ]:


import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import numpy as np
import calcTimeNow
import getDaysInMonth
import getNameNumbers_ec1
finalFile = f'{new_path}{this_month}{date}{thisYear}_TempestT.xlsx'
path1 = '/var/www/html/000/'    
print(finalFile)

df = pd.read_excel(finalFile)
print(df)

t = df['Temperature']
y = t.to_numpy()

plt.figure(figsize= (10,6))
#plt.ylim(60, 105)
plt.yticks(fontsize=12)
plt.ylabel('Temperature (F)', fontsize=12, fontweight ='bold')
#plt.locator_params(axis='y', nbins=20)

#plt.xlim(0,23)
plt.xticks(np.arange(0,24,1))
plt.xlabel('Hours (24 hour clock)', fontsize=12, fontweight ='bold')

plt.title(f'{this_month} {date} {thisYear} Hourly Temperatures', fontsize=12, fontweight ='bold')
plt.plot(y, marker = "*", color = "red", linewidth = 4, label = "Temperature")
plt.grid(True)
plt.autoscale(enable = True, axis = 'y', tight = True)

plt.savefig(f'{path1}testTT')  

