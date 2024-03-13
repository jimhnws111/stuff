#!/usr/bin/env python
# coding: utf-8

# In[1]:


import json
import urllib
import requests
import csv
import re
import calcOneDay

# Calculate the time and date for end of day calculations

xy = calcOneDay.calcOneDay()
start, end = (xy[0], xy[1])


# In[ ]:


import collections
import hashlib
import hmac
import time
from datetime import datetime
import requests
import json
import pprint
import dataFile

parameters = {
  "api-key": "vy8jbrjsxlbwgojepq3vfyfqfywyhvbd", 
  "api-secret": "sdqfm6wdfy9w0pqp2vdka38o6b4vcsvc",
  "station-id": 81211, 
  "end-timestamp": end,
  "start-timestamp": start,
  "t": int(time.time())
}

parameters = collections.OrderedDict(sorted(parameters.items()))

for key in parameters:
    print("Parameter name: \"{}\" has value \"{}\"".format(key, parameters[key]))

apiSecret = parameters["api-secret"];
parameters.pop("api-secret", None);

data = ""
for key in parameters:
    data = data + key + str(parameters[key])

print("Data string to hash is: \"{}\"".format(data))
print('\n')

"""
Calculate the HMAC SHA-256 hash that will be used as the API Signature.
"""
apiSignature = hmac.new(
  apiSecret.encode('utf-8'),
  data.encode('utf-8'),
  hashlib.sha256
).hexdigest()

"""
Let's see what the final API Signature looks like.
"""
print("API Signature is: \"{}\"".format(apiSignature))
print('\n')

# Building the URL to get the station

first_part = ('https://api.weatherlink.com/v2/historic/81211?')
api_key = ('api-key=vy8jbrjsxlbwgojepq3vfyfqfywyhvbd')
add_apisig = ('&api-signature=')
add_t = ('&t='+ str(int(time.time())))

start1 = "&start-timestamp=" + start
end1 = "&end-timestamp=" + end

#
URLfinal = (first_part + api_key + add_t + start1 + end1 + add_apisig + apiSignature)
print(URLfinal)

r =  requests.get(URLfinal)
path_name = '/home/ec2-user/'
dump_file = 'davisDataDump.csv'
data_file = f'{path_name}{dump_file}'

with open(data_file, "w") as fd:   
     json.dump(r.json(), fd)


# In[ ]:


import time
from datetime import datetime
import requests
import json
import dataFile

with open(data_file) as fr:
    davisAPI = json.load(fr) 
    
a = davisAPI['sensors']    
b = a[1]
c = (b['data'])
cLen = len(c)
print(cLen)

with open(data_file, 'w') as outfile: 
    i = 0
    while i < cLen:
        d = c[i]
        hi_temp = (d['temp_hi'])
        lo_temp = (d['temp_lo'])
        rainfall = (d['rainfall_in'])
        dew_hi = (d['dew_point_hi'])
        dew_lo = (d['dew_point_lo'])
        print(hi_temp,lo_temp,rainfall,dew_hi,dew_lo)
        print(f'{hi_temp},{lo_temp},{rainfall},{dew_hi},{dew_lo}',file = outfile)
        i += 1
      
        


# In[ ]:


import pandas as pd
import dataFile
import getPath
import getNameNumbers

#path_name = getPath.getPath_ec2()
#full_file = dataFile.dataFile_ec2()

df = pd.read_csv(data_file, index_col=False,names=['temp_hi', 'temp_lo', 'rainfall', 'dew_hi', 'dew_lo'])

pd.set_option('display.max_rows', 1440)
pd.set_option('display.max_columns', 35)
pd.set_option('display.width', 1500)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 2)

max_temp  = (df.sort_values(by='temp_hi', ascending=False))
max_T = max_temp.iloc[:1]
maxT = max_T['temp_hi'].values[0]
maxT = round(maxT)

min_temp  = (df.sort_values(by='temp_lo', ascending=True))
min_T = min_temp.iloc[:1]
minT = min_T['temp_lo'].values[0]
minT = round(minT)

dew_max = (df.sort_values(by='dew_hi', ascending=False))
dew_max1 = dew_max.iloc[:1]
dewMax = dew_max1['dew_hi'].values[0]
dewMaxT = round(dewMax)

dew_min = (df.sort_values(by='dew_lo', ascending=True))
dew_min1 = dew_min.iloc[:1]
dewMin = dew_min1['dew_lo'].values[0]
dewMinT = round(dewMin)

totR = df['rainfall'].sum()


# In[ ]:


import openpyxl
from openpyxl import load_workbook
import excelFilename
import calcTimeNow
import getNameNumbers

#
# Create the month name for the xlsx filename
#

gg = getNameNumbers.davis()
xls_filename, xls_fullfile, path_name, date, this_month, thisYear = gg[0], gg[1], gg[2], gg[3], gg[4], gg[5]

wb = openpyxl.load_workbook(xls_fullfile)
sheet = wb.active

# Write headers first...
a1 = sheet['A1']
a1.value = "Year"
b1 = sheet['B1']
b1.value = thisYear
c1 = sheet['C1']
c1.value = 'Month'
d1 = sheet['D1']
d1.value = this_month

a3 = sheet['A3']
a3.value = "Date"
b3 = sheet['B3']
b3.value = 'High'
c3 = sheet['C3']
c3.value = 'Low'
d3 = sheet['D3']
d3.value = 'Average'

e3 = sheet['E3']
e3.value = "HDD"
f3 = sheet['F3']
f3.value = 'CDD'
g3 = sheet['G3']
g3.value = 'Rainfall'
h3 = sheet['H3']
h3.value = 'Max_Dew_Pt'
i3 = sheet['I3']
i3.value = 'Min_Dew_Pt'

k4 = sheet['K4']
k4.value = "Highs >=90"
k5 = sheet['K5']
k5.value = "Highs <= 32"
k6 = sheet['K6']
k6.value = 'Lows <= 32'
k7 = sheet['K7']
k7.value = 'Lows <= 0'

k13 = sheet['K14']
k13.value = "Total Rainfall"
k14 = sheet['K15']
k14.value = "rain>=0.01"
k15 = sheet['K16']
k15.value = 'rain>=0.10'
k16 = sheet['K17']
k16.value = 'rain>=0.50'
k17= sheet['K18']
k17.value = 'rain>=1.00'
k23 = sheet['K24']
k23.value = 'Monthly Average'
k24 = sheet['K25']
k24.value = 'Departure'

m3 = sheet['M4']
m3.value = "High"
m4 = sheet['M5']
m4.value = "Low"
m13 = sheet['M14']
m13.value = "Max Rain"
m23 = sheet['M24']
m23.value = "Monthy Rainfall"
m24 = sheet['M25']
m24.value = "Departure"

o3 = sheet['O4']
o3.value = "Date"
o4 = sheet['O5']
o4.value = "Date"


# Calculate the date and write the data...
offset_day = (int(date) + 2)

maxTT = sheet.cell(row = offset_day, column = 2)
maxTT.value = maxT
minTT = sheet.cell(row = offset_day, column = 3)
minTT.value = minT
totRR = sheet.cell(row = offset_day, column = 7)
totRR.value = totR
dewMaxTT = sheet.cell(row = offset_day, column = 8)
dewMaxTT.value = dewMaxT
dewMinTT = sheet.cell(row = offset_day, column = 9)
dewMinTT.value = dewMinT

wb.save(xls_fullfile)


# In[ ]:


import setUpHTML

# Read the Excel file as a possible pandas dataframe and html file

html_path = '/var/www/html/000/'

df1 = pd.read_excel(xls_fullfile, skiprows = 2, names = ['Date','High','Low','Average','HDD','CDD','Rainfall','Max Dew Pt','Min Dew Pt','dead2','dead3','dead4','dead5','dead6','dead7'])
df1 = df1.drop(df1.columns[[9,10,11,12,13,14]], axis = 1)
df1
df1.to_html(f'{html_path}testDavis.html', index = False) 

