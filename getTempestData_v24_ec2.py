#!/usr/bin/env python
# coding: utf-8

# In[11]:


import json
import urllib
import requests
import csv
import re
import calcOneDay

# Calculate the time and date for end of day calculations

xy = calcOneDay.calcOneDay()
start, end = (xy[0], xy[1])


# In[6]:


import datetime
from datetime import datetime
import dataFile

#
# Get data from the Tempest database for the new station
#

token = '877f6425-04a5-4f33-86e7-7123b7ef53d9'
protocol = 'https://'
urlSiteDevice = 'swd.weatherflow.com/swd/rest/observations/device/'
urlSiteStation = 'swd.weatherflow.com/swd/rest/observations/station/'
deviceID = '246921'
stationID = '95775'
preToken = '&token='
preStart = '?time_start='
preEnd = '&time_end='
start_time = start
end_time = end
dayOffset = '&day_offset=1'
format1 = '&format=csv'

#
# Put it together
# 

goGetDeviceSummary = (f'{protocol}{urlSiteDevice}{deviceID}{preStart}{start_time}{preEnd}{end_time}{format1}{preToken}{token}')
print(goGetDeviceSummary)
r =  requests.get(goGetDeviceSummary)
full_file = dataFile.dataFile_wxflow()

with open(full_file,'w') as fd:
     fd.write(r.text)


# In[1]:


import pandas as pd
import dataFile

#
# Read in the CSV file for processing in pandas
#

full_file = dataFile.dataFile_wxflow()
df = pd.read_csv(full_file, index_col=False)

pd.set_option('display.max_rows', 1440)
pd.set_option('display.max_columns', 35)
pd.set_option('display.width', 1500)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 2)

max_temp  = (df.sort_values(by='temperature', ascending=False))
max_T = max_temp.iloc[:1]
maxT = max_T['temperature'].values[0]
maxT = round((maxT*1.8) + 32)

min_temp  = (df.sort_values(by='temperature', ascending=True))
min_T = min_temp.iloc[:1]
minT = min_T['temperature'].values[0]
minT = round((minT*1.8) + 32)

max_pres = (df.sort_values(by='pressure', ascending=False))
max_P = max_pres.iloc[:1]
maxP = max_P['pressure'].values[0]

min_pres = (df.sort_values(by='pressure', ascending=True))
min_P = min_pres.iloc[:1]
minP = min_P['pressure'].values[0]

max_pcpn = (df.sort_values(by='local_daily_precip', ascending=False))
max_rain = max_pcpn.iloc[:1]
maxR = max_rain['local_daily_precip'].values[0]
maxR = round((maxR*0.03937), 2)

tot_rain = df['precip'].sum()
totR = round((tot_rain*0.03937), 2)

strike_distance = (df.sort_values(by='strike_distance', ascending=False))
lightning = (df['strike_distance'] > 0 & (df['strike_distance'] < 16).all())
df.insert(8,'lightning',lightning)

# Determine if the strike distance is close enough to count as a thunderstorm
x = len(df)
a = 0

while a < x:
        if (df['lightning'] == True).any():
            q = "Yes"
        else:
            q = "No"
        a += 1
           
strike_count = df['strike_count'].sum()


# In[2]:


#
# Write to the appropriate Excel file
#

import openpyxl
from openpyxl import load_workbook
import datetime
from datetime import datetime
import excelFilename
import calcTimeNow
from calcTimeNow import calcTimeNow

#
# Create the month name for the xlsx filename
#

sa1 = excelFilename.tempest_ec2()
xls_filename, path_name = (sa1[0], sa1[1])
xls_fullfile = (f'{path_name}{xls_filename}')
t1 = calcTimeNow()
now, myMonth, myYear, date = (t1[0], t1[1], t1[2], t1[3])
wb = openpyxl.load_workbook(xls_fullfile)
sheet = wb.active

# Write headers first...
a1 = sheet['A1']
a1.value = "Year"
b1 = sheet['B1']
b1.value = myYear
c1 = sheet['C1']
c1.value = 'Month'
d1 = sheet['D1']
d1.value = myMonth

a3 = sheet['A3']
a3.value = "Date"
b3 = sheet['B3']
b3.value = 'High'
c3 = sheet['C3']
c3.value = 'Low'
d3 = sheet['D3']
d3.value = 'Average'

e3 = sheet['E3']
e3.value = "HDD"
f3 = sheet['F3']
f3.value = 'CDD'
g3 = sheet['G3']
g3.value = 'totR'
h3 = sheet['H3']
h3.value = 'lightning'

k3 = sheet['K3']
k3.value = "Highs >=90"
k4 = sheet['K4']
k4.value = "Highs <= 32"
k5 = sheet['K5']
k5.value = 'Lows <= 32'
k6 = sheet['K6']
k6.value = 'Lows <= 0'

k13 = sheet['K13']
k13.value = "Total Rainfall"
k14 = sheet['K14']
k14.value = "rain>=0.01"
k15 = sheet['K15']
k15.value = 'rain>=0.01'
k16 = sheet['K16']
k16.value = 'rain>=0.50'
k17= sheet['K17']
k17.value = 'rain>=1.00'
k23 = sheet['K23']
k23.value = 'Monthly Average'
k24 = sheet['K24']
k24.value = 'Departure'

m3 = sheet['M3']
m3.value = "High"
m4 = sheet['M4']
m4.value = "Low"
m13 = sheet['M13']
m13.value = "Max Rain"
m23 = sheet['M23']
m23.value = "Monthy Rainfall"
m24 = sheet['M24']
m24.value = "Departure"

o3 = sheet['O3']
o3.value = "Date"
o4 = sheet['O4']
o4.value = "Date"

# Write the data..

offset_day = (int(date) + 2)
maxTT = sheet.cell(row = offset_day, column = 2)
maxTT.value = maxT
minTT = sheet.cell(row = offset_day, column = 3)
minTT.value = minT
totRR = sheet.cell(row = offset_day, column = 7)
totRR.value = totR
lightning1 = sheet.cell(row = offset_day, column = 8)
lightning1.value = q

wb.save(xls_fullfile)


# In[10]:


import pandas as pd
import setUpHTML

# Read the Excel file as a possible pandas dataframe and html file

t2 = setUpHTML.setUpHTML_tempest_ec2()
new_file, html_path = (t2[0], t2[1])

df1 = pd.read_excel(new_file, skiprows = 2, names = ['Date','High','Low','Average','HDD','CDD','totR','Lightning?', 'dead1','dead2','dead3','dead4','dead5','dead6','dead7'])
df1 = df1.drop(df1.columns[[8,9,10,11,12,13,14]], axis = 1)
df1.to_html(f'{html_path}testTempest.html', index = False) 


# In[ ]:




